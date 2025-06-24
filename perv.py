import sqlite3
from datetime import datetime

# Bazaga ulanish
conn = sqlite3.connect("db.sqlite3")
cursor = conn.cursor()

# Barcha sanalarni o'qib, matn formatida yozib chiqamiz
cursor.execute("SELECT id, aniqlangan_sanasi FROM ish_excelupload WHERE aniqlangan_sanasi IS NOT NULL")

rows = cursor.fetchall()
for row in rows:
    row_id = row[0]
    original_date = row[1]  

    # Formatlash: datetime → 'dd.mm.yyyy'
    dt = datetime.strptime(original_date, "%Y-%m-%d %H:%M:%S")
    #dt = datetime.strptime(original_date, "%d.%m.%Y")
    formatted_date = dt.strftime("%d.%m.%Y")

    # Yangilash: formatlangan matnni yozish
    cursor.execute("""
        UPDATE ish_excelupload
        SET aniqlangan_sanasi = ?
        WHERE id = ?
    """, (formatted_date, row_id))

# O'zgarishlarni saqlash
conn.commit()
conn.close()


exit()

import sqlite3

# Bazaga ulanish
conn = sqlite3.connect("db.sqlite3")
cursor = conn.cursor()

# NULL qiymatlarni 0 ga almashtirish
cursor.execute("""
    UPDATE ish_excelupload
    SET faoliyatsiz = 0
    WHERE faoliyatsiz = ''
""")

# O'zgarishlarni saqlash va yopish

# O'zgarishlar sonini olish
ozgarishlar_soni = cursor.rowcount

# Natijani ko‘rsatish
print(f"{ozgarishlar_soni} ta qator yangilandi.")
conn.commit()
conn.close()




exit()

import sqlite3
from openpyxl import load_workbook


# Load Excel workbook and sheet
wb = load_workbook("perv.xlsx")
sheet = wb.active  # You can use wb['SheetName'] if needed

# Connect to SQLite database (or create it)
conn = sqlite3.connect("db.sqlite3")
cursor = conn.cursor()

# Create table
#cursor.execute('''
#    CREATE TABLE IF NOT EXISTS ish_excelupload (
#        id INTEGER PRIMARY KEY,
#        name TEXT,
#        age INTEGER
#    )
#''')
#
# Read Excel rows and insert into SQLite
rows = list(sheet.iter_rows(min_row=2, values_only=True))  # Skip header
for row in rows:
    cursor.execute("INSERT INTO ish_excelupload (okpo, inn, nomi, soato, hisobot_nomi, hisobot_davri, aniqlangan_sanasi, sababi, xat_turi, xat_sanasi, nazoratdan_chiqarilgan, izoh, sudga_chiqarilgan, dalolatnomasi_mavjudligi, faoliyatsiz, opf, tasdiqlangan) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", row)
conn.commit()
conn.close()